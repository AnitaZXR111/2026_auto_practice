import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment


input_path = "../data/sales_data.xlsx"
output_path = "../output/格式调整.xlsx"

wb = openpyxl.load_workbook(input_path)
df = wb["Sheet1"]


for col in range(1, df.max_column + 1):
    max_length = 0
    for row in range(1, df.max_row + 1):
        cell = df.cell(row=row, column=col)
        if cell.value:
            length = len(str(cell.value))
        else:
            length = 0
        
        if length > max_length:
            max_length = length
    
    adjusted_width = min(max_length + 3, 12)
    df.column_dimensions[get_column_letter(col)].width = adjusted_width

    
    cell = df.cell(row = 1, column = col)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')



for r in range(2,df.max_row + 1):
    cell = df.cell(row = r, column = 4)
    if cell.value is not None and cell.value<10:
        cell.font = Font(color='FF0000')

wb.save(output_path)
print("="*50)
print(f"已保存至: {output_path}")
print("="*50)
